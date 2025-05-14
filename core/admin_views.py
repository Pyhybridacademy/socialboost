from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Q, Count, F
from django.utils import timezone
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.core.management import call_command
from django.conf import settings
from decimal import Decimal
import csv
import io
import datetime
from .models import Service, Order, Transaction, UserBalance, SiteSettings, ProviderPayment
from .api_client import SocialMediaAPI
import logging

logger = logging.getLogger(__name__)

# Admin Users Management
@staff_member_required
def admin_users(request):
    # Get all users
    users_list = User.objects.all()
    
    # Handle search
    search_query = request.GET.get('search', '')
    if search_query:
        users_list = users_list.filter(
            Q(username__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # Handle status filter
    status = request.GET.get('status', 'all')
    if status == 'active':
        users_list = users_list.filter(is_active=True)
    elif status == 'inactive':
        users_list = users_list.filter(is_active=False)
    elif status == 'staff':
        users_list = users_list.filter(is_staff=True)
    
    # Handle sorting
    sort_by = request.GET.get('sort', 'newest')
    if sort_by == 'newest':
        users_list = users_list.order_by('-date_joined')
    elif sort_by == 'oldest':
        users_list = users_list.order_by('date_joined')
    elif sort_by == 'balance_high':
        users_list = users_list.order_by('-userbalance__balance')
    elif sort_by == 'balance_low':
        users_list = users_list.order_by('userbalance__balance')
    elif sort_by == 'orders_high':
        users_list = users_list.annotate(order_count=Count('order')).order_by('-order_count')
    
    # Annotate with order count
    users_list = users_list.annotate(order_count=Count('order'))
    
    # Pagination
    paginator = Paginator(users_list, 20)  # Show 20 users per page
    page = request.GET.get('page', 1)
    
    try:
        users = paginator.page(page)
    except:
        users = paginator.page(1)
    
    context = {
        'users': users,
        'search_query': search_query,
        'status': status,
        'sort_by': sort_by,
    }
    
    return render(request, 'core/admin/users.html', context)

@staff_member_required
def admin_user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user_balance = UserBalance.objects.get_or_create(user=user)[0]
    
    # Get user statistics
    total_orders = Order.objects.filter(user=user).count()
    completed_orders = Order.objects.filter(user=user, status='completed').count()
    pending_orders = Order.objects.filter(user=user, status='pending').count()
    total_spent = Order.objects.filter(user=user).aggregate(Sum('charge'))['charge__sum'] or 0
    
    # Get total deposits
    total_deposits = Transaction.objects.filter(
        user=user, 
        transaction_type='deposit'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Get recent orders
    recent_orders = Order.objects.filter(user=user).order_by('-created_at')[:5]
    
    # Get recent transactions
    recent_transactions = Transaction.objects.filter(user=user).order_by('-created_at')[:5]
    
    # Get last activity
    last_activity = recent_transactions.first().created_at if recent_transactions.exists() else None
    if not last_activity and recent_orders.exists():
        last_activity = recent_orders.first().created_at
    
    context = {
        'user': user,
        'user_balance': user_balance,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'total_spent': total_spent,
        'total_deposits': total_deposits,
        'recent_orders': recent_orders,
        'recent_transactions': recent_transactions,
        'last_activity': last_activity,
    }
    
    return render(request, 'core/admin/user_detail.html', context)

@staff_member_required
def admin_user_edit(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        user = get_object_or_404(User, id=user_id)
        
        user.email = request.POST.get('email', user.email)
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.is_active = 'is_active' in request.POST
        user.is_staff = 'is_staff' in request.POST
        
        user.save()
        
        # Check if balance was updated
        if 'balance' in request.POST and request.POST.get('balance'):
            try:
                new_balance = Decimal(request.POST.get('balance'))
                user_balance = UserBalance.objects.get_or_create(user=user)[0]
                
                # Create a transaction record for the balance adjustment
                if new_balance != user_balance.balance:
                    adjustment = new_balance - user_balance.balance
                    transaction_type = 'deposit' if adjustment > 0 else 'withdrawal'
                    
                    Transaction.objects.create(
                        user=user,
                        amount=abs(adjustment),
                        transaction_type=transaction_type,
                        description=f"Admin balance adjustment by {request.user.username}"
                    )
                    
                    user_balance.balance = new_balance
                    user_balance.save()
            except:
                messages.error(request, "Invalid balance amount")
        
        messages.success(request, f"User {user.username} updated successfully")
        
        # Redirect back to the user's detail page if we came from there
        referer = request.META.get('HTTP_REFERER', '')
        if f'/admin/users/{user_id}/' in referer:
            return redirect('admin_user_detail', user_id=user_id)
        
        return redirect('admin_users')
    
    return redirect('admin_users')

@staff_member_required
def admin_adjust_balance(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user_balance = UserBalance.objects.get_or_create(user=user)[0]
    
    if request.method == 'POST':
        adjustment_type = request.POST.get('adjustment_type')
        amount_str = request.POST.get('amount', '0')
        reason = request.POST.get('reason', '')
        
        try:
            amount = Decimal(amount_str)
            if amount < 0:
                raise ValueError("Amount cannot be negative")
            
            current_balance = user_balance.balance
            new_balance = current_balance
            
            if adjustment_type == 'add':
                new_balance = current_balance + amount
                transaction_type = 'deposit'
                transaction_amount = amount
            elif adjustment_type == 'subtract':
                if amount > current_balance:
                    messages.error(request, f"Cannot subtract ₦{amount} from balance of ₦{current_balance}")
                    return redirect('admin_user_detail', user_id=user_id)
                new_balance = current_balance - amount
                transaction_type = 'withdrawal'
                transaction_amount = amount
            elif adjustment_type == 'set':
                if amount > current_balance:
                    transaction_type = 'deposit'
                    transaction_amount = amount - current_balance
                else:
                    transaction_type = 'withdrawal'
                    transaction_amount = current_balance - amount
                new_balance = amount
            
            # Create transaction record
            if transaction_amount > 0:
                Transaction.objects.create(
                    user=user,
                    amount=transaction_amount,
                    transaction_type=transaction_type,
                    description=f"Admin balance adjustment: {reason}"
                )
            
            # Update balance
            user_balance.balance = new_balance
            user_balance.save()
            
            messages.success(request, f"Balance for {user.username} updated to ₦{new_balance}")
            
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Invalid amount: {str(e)}")
        
        return redirect('admin_user_detail', user_id=user_id)
    
    return redirect('admin_user_detail', user_id=user_id)

@staff_member_required
def admin_user_deactivate(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        
        user.is_active = False
        user.save()
        
        messages.success(request, f"User {user.username} has been deactivated")
        logger.info(f"User {user.username} deactivated by admin {request.user.username}. Reason: {reason}")
        
        return redirect('admin_user_detail', user_id=user_id)
    
    return redirect('admin_user_detail', user_id=user_id)

@staff_member_required
def admin_user_activate(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        user.is_active = True
        user.save()
        
        messages.success(request, f"User {user.username} has been activated")
        logger.info(f"User {user.username} activated by admin {request.user.username}")
        
        return redirect('admin_user_detail', user_id=user_id)
    
    return redirect('admin_user_detail', user_id=user_id)

@staff_member_required
def admin_user_orders(request, user_id):
    user = get_object_or_404(User, id=user_id)
    orders = Order.objects.filter(user=user).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(orders, 20)
    page = request.GET.get('page', 1)
    
    try:
        orders_page = paginator.page(page)
    except:
        orders_page = paginator.page(1)
    
    context = {
        'user': user,
        'orders': orders_page,
    }
    
    return render(request, 'core/admin/user_orders.html', context)

@staff_member_required
def admin_user_transactions(request, user_id):
    user = get_object_or_404(User, id=user_id)
    transactions = Transaction.objects.filter(user=user).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page', 1)
    
    try:
        transactions_page = paginator.page(page)
    except:
        transactions_page = paginator.page(1)
    
    context = {
        'user': user,
        'transactions': transactions_page,
    }
    
    return render(request, 'core/admin/user_transactions.html', context)

@staff_member_required
def admin_export_users(request):
    format_type = request.GET.get('format', 'csv')
    status = request.GET.get('status', 'all')
    
    # Filter users based on status
    if status == 'active':
        users = User.objects.filter(is_active=True)
    elif status == 'inactive':
        users = User.objects.filter(is_active=False)
    else:
        users = User.objects.all()
    
    # Create CSV or Excel file
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined', 'Last Login', 'Active', 'Staff', 'Balance', 'Orders Count'])
        
        for user in users:
            try:
                balance = user.userbalance.balance
            except UserBalance.DoesNotExist:
                balance = 0
            
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.date_joined,
                user.last_login,
                user.is_active,
                user.is_staff,
                balance,
                Order.objects.filter(user=user).count()
            ])
        
        return response
    else:
        # For Excel export (requires openpyxl or xlwt)
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"
            
            # Add headers
            headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date Joined', 'Last Login', 'Active', 'Staff', 'Balance', 'Orders Count']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            
            # Add data
            row_num = 2
            for user in users:
                try:
                    balance = user.userbalance.balance
                except UserBalance.DoesNotExist:
                    balance = 0
                
                ws.cell(row=row_num, column=1).value = user.id
                ws.cell(row=row_num, column=2).value = user.username
                ws.cell(row=row_num, column=3).value = user.email
                ws.cell(row=row_num, column=4).value = user.first_name
                ws.cell(row=row_num, column=5).value = user.last_name
                ws.cell(row=row_num, column=6).value = user.date_joined
                ws.cell(row=row_num, column=7).value = user.last_login
                ws.cell(row=row_num, column=8).value = user.is_active
                ws.cell(row=row_num, column=9).value = user.is_staff
                ws.cell(row=row_num, column=10).value = float(balance)
                ws.cell(row=row_num, column=11).value = Order.objects.filter(user=user).count()
                
                row_num += 1
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="users_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
        
        except ImportError:
            messages.error(request, "Excel export requires openpyxl library")
            return redirect('admin_users')

# Admin Services Management
@staff_member_required
def admin_services(request):
    # Get all services
    services_list = Service.objects.all()
    
    # Handle search
    search_query = request.GET.get('search', '')
    if search_query:
        services_list = services_list.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(category__icontains=search_query)
        )
    
    # Handle category filter
    selected_category = request.GET.get('category', '')
    if selected_category:
        services_list = services_list.filter(category=selected_category)
    
    # Handle status filter
    status = request.GET.get('status', 'all')
    if status == 'active':
        services_list = services_list.filter(is_active=True)
    elif status == 'inactive':
        services_list = services_list.filter(is_active=False)
    
    # Handle sorting
    sort_by = request.GET.get('sort', 'newest')
    if sort_by == 'newest':
        services_list = services_list.order_by('-id')
    elif sort_by == 'oldest':
        services_list = services_list.order_by('id')
    elif sort_by == 'name':
        services_list = services_list.order_by('name')
    elif sort_by == 'rate_high':
        services_list = services_list.order_by('-rate')
    elif sort_by == 'rate_low':
        services_list = services_list.order_by('rate')
    
    # Get all unique categories
    categories = Service.objects.values_list('category', flat=True).distinct()
    
    # Pagination
    paginator = Paginator(services_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        services = paginator.page(page)
    except:
        services = paginator.page(1)
    
    context = {
        'services': services,
        'categories': categories,
        'selected_category': selected_category,
        'search_query': search_query,
        'status': status,
        'sort_by': sort_by,
    }
    
    return render(request, 'core/admin/services.html', context)

@staff_member_required
def admin_service_detail(request, service_id):
    service = get_object_or_404(Service, id=service_id)
    
    # Get orders for this service
    orders = Order.objects.filter(service=service).order_by('-created_at')[:10]
    
    # Calculate order statistics
    total_orders = Order.objects.filter(service=service).count()
    completed_orders = Order.objects.filter(service=service, status='completed').count()
    pending_orders = Order.objects.filter(service=service, status='pending').count()
    processing_orders = Order.objects.filter(service=service, status='processing').count()
    
    # Calculate revenue
    total_revenue = Order.objects.filter(service=service).aggregate(Sum('charge'))['charge__sum'] or 0
    
    # Get site settings for markup calculation
    site_settings = SiteSettings.get_settings()
    markup_multiplier = (Decimal('100') + site_settings.markup_percentage) / Decimal('100')
    
    # Calculate wholesale cost (original price before markup)
    wholesale_rate = (service.rate / markup_multiplier).quantize(Decimal('0.01'))
    
    context = {
        'service': service,
        'orders': orders,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'total_revenue': total_revenue,
        'wholesale_rate': wholesale_rate,
        'markup_percentage': site_settings.markup_percentage,
    }
    
    return render(request, 'core/admin/service_detail.html', context)

@staff_member_required
def admin_service_edit(request):
    if request.method == 'POST':
        service_id = request.POST.get('service_id')
        service = get_object_or_404(Service, id=service_id)
        
        service.name = request.POST.get('name', service.name)
        service.category = request.POST.get('category', service.category)
        
        try:
            service.rate = Decimal(request.POST.get('rate', service.rate))
        except:
            messages.error(request, "Invalid rate amount")
        
        try:
            service.min_quantity = int(request.POST.get('min_quantity', service.min_quantity))
            service.max_quantity = int(request.POST.get('max_quantity', service.max_quantity))
        except:
            messages.error(request, "Invalid quantity values")
        
        service.description = request.POST.get('description', service.description)
        service.is_active = 'is_active' in request.POST
        
        service.save()
        
        messages.success(request, f"Service '{service.name}' updated successfully")
        
        # Redirect back to the service's detail page if we came from there
        referer = request.META.get('HTTP_REFERER', '')
        if f'/admin/services/{service_id}/' in referer:
            return redirect('admin_service_detail', service_id=service_id)
        
        return redirect('admin_services')
    
    return redirect('admin_services')

@staff_member_required
def admin_service_toggle(request):
    if request.method == 'POST':
        service_id = request.POST.get('service_id')
        is_active = request.POST.get('is_active') == 'true'
        
        service = get_object_or_404(Service, id=service_id)
        service.is_active = is_active
        service.save()
        
        status_text = "activated" if is_active else "deactivated"
        messages.success(request, f"Service '{service.name}' has been {status_text}")
        
        return redirect('admin_services')
    
    return redirect('admin_services')

@staff_member_required
def admin_export_services(request):
    format_type = request.GET.get('format', 'csv')
    status = request.GET.get('status', 'all')
    category = request.GET.get('category', '')
    
    # Filter services
    services = Service.objects.all()
    
    if status == 'active':
        services = services.filter(is_active=True)
    elif status == 'inactive':
        services = services.filter(is_active=False)
    
    if category:
        services = services.filter(category=category)
    
    # Create CSV or Excel file
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="services_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Service ID', 'Name', 'Category', 'Rate', 'Min Quantity', 'Max Quantity', 'Active', 'Description'])
        
        for service in services:
            writer.writerow([
                service.id,
                service.service_id,
                service.name,
                service.category,
                float(service.rate),
                service.min_quantity,
                service.max_quantity,
                service.is_active,
                service.description
            ])
        
        return response
    else:
        # For Excel export (requires openpyxl or xlwt)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Services"
            
            # Add headers
            headers = ['ID', 'Service ID', 'Name', 'Category', 'Rate', 'Min Quantity', 'Max Quantity', 'Active', 'Description']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            
            # Add data
            row_num = 2
            for service in services:
                ws.cell(row=row_num, column=1).value = service.id
                ws.cell(row=row_num, column=2).value = service.service_id
                ws.cell(row=row_num, column=3).value = service.name
                ws.cell(row=row_num, column=4).value = service.category
                ws.cell(row=row_num, column=5).value = float(service.rate)
                ws.cell(row=row_num, column=6).value = service.min_quantity
                ws.cell(row=row_num, column=7).value = service.max_quantity
                ws.cell(row=row_num, column=8).value = service.is_active
                ws.cell(row=row_num, column=9).value = service.description
                
                row_num += 1
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="services_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
        
        except ImportError:
            messages.error(request, "Excel export requires openpyxl library")
            return redirect('admin_services')

# Admin Orders Management
@staff_member_required
def admin_orders(request):
    # Get all orders
    orders_list = Order.objects.all()
    
    # Handle search
    search_query = request.GET.get('search', '')
    if search_query:
        # Try to convert search query to integer for order ID search
        try:
            order_id = int(search_query)
            orders_list = orders_list.filter(
                Q(id=order_id) |
                Q(user__username__icontains=search_query) |
                Q(service__name__icontains=search_query)
            )
        except ValueError:
            orders_list = orders_list.filter(
                Q(user__username__icontains=search_query) |
                Q(service__name__icontains=search_query)
            )
    
    # Handle status filter
    status = request.GET.get('status', 'all')
    if status != 'all':
        orders_list = orders_list.filter(status=status)
    
    # Handle date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        try:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            orders_list = orders_list.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            orders_list = orders_list.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Handle sorting
    sort_by = request.GET.get('sort', 'newest')
    if sort_by == 'newest':
        orders_list = orders_list.order_by('-created_at')
    elif sort_by == 'oldest':
        orders_list = orders_list.order_by('created_at')
    elif sort_by == 'amount_high':
        orders_list = orders_list.order_by('-charge')
    elif sort_by == 'amount_low':
        orders_list = orders_list.order_by('charge')
    
    # Pagination
    paginator = Paginator(orders_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        orders = paginator.page(page)
    except:
        orders = paginator.page(1)
    
    context = {
        'orders': orders,
        'search_query': search_query,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
    }
    
    return render(request, 'core/admin/orders.html', context)

@staff_member_required
def admin_order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    # If the order is still processing, check for updates
    if order.status == 'processing' and order.order_id:
        api = SocialMediaAPI(settings.API_KEY)
        response = api.get_order_status(order.order_id)
        
        if 'status' in response:
            # Map API status to our status
            status_mapping = {
                'Pending': 'pending',
                'In progress': 'processing',
                'Completed': 'completed',
                'Canceled': 'canceled',
                'Error': 'error'
            }
            
            new_status = status_mapping.get(response['status'], order.status)
            
            if new_status != order.status:
                order.status = new_status
                order.save()
    
    # Get site settings for markup calculation
    site_settings = SiteSettings.get_settings()
    markup_multiplier = (Decimal('100') + site_settings.markup_percentage) / Decimal('100')
    
    # Calculate wholesale cost (original price before markup)
    wholesale_cost = (order.charge / markup_multiplier).quantize(Decimal('0.01'))
    
    # Calculate profit
    profit = order.charge - wholesale_cost
    
    # Get order status history (if you have a model for that)
    # status_history = OrderStatusHistory.objects.filter(order=order).order_by('created_at')
    
    context = {
        'order': order,
        'wholesale_cost': wholesale_cost,
        'profit': profit,
        'markup_percentage': site_settings.markup_percentage,
        # 'status_history': status_history,
    }
    
    return render(request, 'core/admin/order_detail.html', context)

@staff_member_required
def admin_order_update_status(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        note = request.POST.get('note', '')
        refund_user = 'refund_user' in request.POST
        
        order = get_object_or_404(Order, id=order_id)
        old_status = order.status
        
        # Update the order status
        order.status = new_status
        order.save()
        
        # Log the status change
        logger.info(f"Order #{order.id} status changed from {old_status} to {new_status} by admin {request.user.username}. Note: {note}")
        
        # Handle refund if requested and status is canceled or error
        if refund_user and new_status in ['canceled', 'error'] and old_status in ['pending', 'processing']:
            user_balance = UserBalance.objects.get(user=order.user)
            user_balance.balance += order.charge
            user_balance.save()
            
            # Create refund transaction
            Transaction.objects.create(
                user=order.user,
                amount=order.charge,
                transaction_type='refund',
                description=f"Refund for order #{order.id} - Status changed to {new_status} by admin"
            )
            
            messages.success(request, f"Order status updated and ₦{order.charge} refunded to user")
        else:
            messages.success(request, "Order status updated successfully")
        
        return redirect('admin_order_detail', order_id=order.id)
    
    return redirect('admin_orders')

@staff_member_required
def admin_export_orders(request):
    format_type = request.GET.get('format', 'csv')
    status = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Filter orders
    orders = Order.objects.all()
    
    if status != 'all':
        orders = orders.filter(status=status)
    
    if date_from:
        try:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Order by creation date
    orders = orders.order_by('-created_at')
    
    # Create CSV or Excel file
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="orders_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'User', 'Service', 'Service ID', 'Link', 'Quantity', 'Charge', 'Status', 'Date'])
        
        for order in orders:
            writer.writerow([
                order.id,
                order.user.username,
                order.service.name,
                order.service.service_id,
                order.link,
                order.quantity,
                float(order.charge),
                order.status,
                order.created_at
            ])
        
        return response
    else:
        # For Excel export (requires openpyxl or xlwt)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Orders"
            
            # Add headers
            headers = ['ID', 'User', 'Service', 'Service ID', 'Link', 'Quantity', 'Charge', 'Status', 'Date']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            
            # Add data
            row_num = 2
            for order in orders:
                ws.cell(row=row_num, column=1).value = order.id
                ws.cell(row=row_num, column=2).value = order.user.username
                ws.cell(row=row_num, column=3).value = order.service.name
                ws.cell(row=row_num, column=4).value = order.service.service_id
                ws.cell(row=row_num, column=5).value = order.link
                ws.cell(row=row_num, column=6).value = order.quantity
                ws.cell(row=row_num, column=7).value = float(order.charge)
                ws.cell(row=row_num, column=8).value = order.status
                ws.cell(row=row_num, column=9).value = order.created_at
                
                row_num += 1
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="orders_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
        
        except ImportError:
            messages.error(request, "Excel export requires openpyxl library")
            return redirect('admin_orders')

# Admin Transactions Management
@staff_member_required
def admin_transactions(request):
    # Get all transactions
    transactions_list = Transaction.objects.all()
    
    # Handle search
    search_query = request.GET.get('search', '')
    if search_query:
        transactions_list = transactions_list.filter(
            Q(user__username__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Handle type filter
    transaction_type = request.GET.get('type', 'all')
    if transaction_type != 'all':
        transactions_list = transactions_list.filter(transaction_type=transaction_type)
    
    # Handle date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        try:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            transactions_list = transactions_list.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            transactions_list = transactions_list.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Handle sorting
    sort_by = request.GET.get('sort', 'newest')
    if sort_by == 'newest':
        transactions_list = transactions_list.order_by('-created_at')
    elif sort_by == 'oldest':
        transactions_list = transactions_list.order_by('created_at')
    elif sort_by == 'amount_high':
        transactions_list = transactions_list.order_by('-amount')
    elif sort_by == 'amount_low':
        transactions_list = transactions_list.order_by('amount')
    
    # Pagination
    paginator = Paginator(transactions_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        transactions = paginator.page(page)
    except:
        transactions = paginator.page(1)
    
    # Get transaction type counts
    transaction_types = Transaction.objects.values('transaction_type').annotate(
        count=Count('transaction_type'),
        total=Sum('amount')
    )
    
    # Convert to dictionary for easier template access
    type_stats = {}
    for t in transaction_types:
        type_stats[t['transaction_type']] = {
            'count': t['count'],
            'total': t['total']
        }
    
    context = {
        'transactions': transactions,
        'search_query': search_query,
        'transaction_type': transaction_type,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'type_stats': type_stats,
    }
    
    return render(request, 'core/admin/transactions.html', context)

@staff_member_required
def admin_export_transactions(request):
    format_type = request.GET.get('format', 'csv')
    transaction_type = request.GET.get('type', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Filter transactions
    transactions = Transaction.objects.all()
    
    if transaction_type != 'all':
        transactions = transactions.filter(transaction_type=transaction_type)
    
    if date_from:
        try:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            transactions = transactions.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            transactions = transactions.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Order by creation date
    transactions = transactions.order_by('-created_at')
    
    # Create CSV or Excel file
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'User', 'Type', 'Amount', 'Description', 'Date'])
        
        for transaction in transactions:
            writer.writerow([
                transaction.id,
                transaction.user.username,
                transaction.transaction_type,
                float(transaction.amount),
                transaction.description,
                transaction.created_at
            ])
        
        return response
    else:
        # For Excel export (requires openpyxl or xlwt)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            # Add headers
            headers = ['ID', 'User', 'Type', 'Amount', 'Description', 'Date']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            
            # Add data
            row_num = 2
            for transaction in transactions:
                ws.cell(row=row_num, column=1).value = transaction.id
                ws.cell(row=row_num, column=2).value = transaction.user.username
                ws.cell(row=row_num, column=3).value = transaction.transaction_type
                ws.cell(row=row_num, column=4).value = float(transaction.amount)
                ws.cell(row=row_num, column=5).value = transaction.description
                ws.cell(row=row_num, column=6).value = transaction.created_at
                
                row_num += 1
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
        
        except ImportError:
            messages.error(request, "Excel export requires openpyxl library")
            return redirect('admin_transactions')

# Admin Site Settings
@staff_member_required
def admin_site_settings(request):
    settings = SiteSettings.get_settings()
    
    if request.method == 'POST':
        # Update markup percentage
        try:
            markup_percentage = Decimal(request.POST.get('markup_percentage', settings.markup_percentage))
            settings.markup_percentage = markup_percentage
        except:
            messages.error(request, "Invalid markup percentage")
        
        # Update minimum withdrawal amount
        try:
            min_withdrawal = Decimal(request.POST.get('min_withdrawal_amount', settings.min_withdrawal_amount))
            settings.min_withdrawal_amount = min_withdrawal
        except:
            messages.error(request, "Invalid minimum withdrawal amount")
        
        # Update support email
        settings.support_email = request.POST.get('support_email', settings.support_email)
        
        # Update maintenance mode
        settings.maintenance_mode = 'maintenance_mode' in request.POST
        
        # Save settings
        settings.save()
        
        messages.success(request, "Site settings updated successfully")
        
        # Check if prices should be updated
        if 'update_prices' in request.POST:
            # Trigger service sync to update prices with new markup
            from django.core.management import call_command
            call_command('sync_services')
            messages.info(request, "Service prices have been updated with the new markup")
        
        return redirect('admin_site_settings')
    
    context = {
        'settings': settings,
    }
    
    return render(request, 'core/admin/site_settings.html', context)

# Manual Command Execution
@staff_member_required
def sync_services_manual(request):
    if request.method == 'POST':
        update_prices = 'update_prices' in request.POST
        
        try:
            # Capture the command output
            from io import StringIO
            import sys
            
            # Redirect stdout to capture output
            old_stdout = sys.stdout
            sys.stdout = mystdout = StringIO()
            
            # Run the command
            from django.core.management import call_command
            call_command('sync_services')
            
            # Get the output
            sys.stdout = old_stdout
            output = mystdout.getvalue()
            
            # Log the output
            logger.info(f"Manual sync_services executed by {request.user.username}. Output: {output}")
            
            messages.success(request, "Services synchronized successfully!")
            
            # Set success message with details
            if 'Service sync complete' in output:
                # Try to extract counts from output
                try:
                    import re
                    counts = re.search(r'Service sync complete: (\d+) created, (\d+) updated, (\d+) deactivated', output)
                    if counts:
                        created, updated, deactivated = counts.groups()
                        messages.info(request, f"{created} services created, {updated} updated, and {deactivated} deactivated")
                except:
                    pass
            
        except Exception as e:
            logger.error(f"Error in manual sync_services: {str(e)}")
            messages.error(request, f"Error synchronizing services: {str(e)}")
        
        return redirect('admin_services')
    
    return redirect('admin_dashboard')

@staff_member_required
def update_orders_manual(request):
    if request.method == 'POST':
        update_all = 'update_all' in request.POST
        
        try:
            # Capture the command output
            from io import StringIO
            import sys
            
            # Redirect stdout to capture output
            old_stdout = sys.stdout
            sys.stdout = mystdout = StringIO()
            
            # Run the command
            from django.core.management import call_command
            call_command('update_order_statuses')
            
            # Get the output
            sys.stdout = old_stdout
            output = mystdout.getvalue()
            
            # Log the output
            logger.info(f"Manual update_order_statuses executed by {request.user.username}. Output: {output}")
            
            messages.success(request, "Order statuses updated successfully!")
            
            # Set success message with details
            if 'Updated' in output:
                # Try to extract count from output
                try:
                    import re
                    count = re.search(r'Updated (\d+) order statuses', output)
                    if count:
                        updated = count.group(1)
                        messages.info(request, f"{updated} order statuses were updated")
                except:
                    pass
            
        except Exception as e:
            logger.error(f"Error in manual update_order_statuses: {str(e)}")
            messages.error(request, f"Error updating order statuses: {str(e)}")
        
        return redirect('admin_orders')
    
    return redirect('admin_dashboard')
